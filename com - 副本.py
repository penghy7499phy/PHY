import math
import xlrd
import numpy as np
import openpyxl

def bend(E,R0,r0,b,h,bi,f,u):   #Ej是抗弯刚度，R0是弯头半径，r0是钢管内半径，ri是软管外半径,t是厚度,b是角度，f是入口拉力，u是摩擦系数
    global Ej,Ri,C,x,Ej
    if bi>180:
        b=180-(360-bi)
    else:
        b=180-bi
    Ej=E*b*h**3/12
    Ej=Ej
    beta=b/180*math.pi
    Ri=R0-r0+h+2*(r0-h)/(1-math.cos(beta))
    C=2*Ej/Ri**2*math.sin(beta/2)
    x=C/(math.sin(beta/2)-beta/2*math.cos(beta/2))
    if f>=x:
        ff=(1+u*math.sin(beta/2))/(1-u*math.sin(beta/2))*f
    if f<x:
        Fc=(C-(math.sin(beta/2)-beta/2*math.cos(beta))*f)/((beta/2+u)*math.sin(beta/2)-u*beta/2*math.cos(beta/2))
        ff=((1+u*math.sin(beta/2))*f+(1+math.cos(beta/2)+u*math.sin(beta/2))*u*Fc)/(1-u*math.sin(beta/2))
        if ff>=x:
            ff=ff
        else:
            z1=1+u*math.sin(beta/2)
            z2=(1+math.cos(beta/2)-u*math.sin(beta/2))*u
            z3=(math.sin(beta/2)-beta/2*math.cos(beta/2))
            z4=(beta/2-u)*math.sin(beta/2)+u*beta/2*math.cos(beta/2)
            z5=(1+math.cos(beta/2)+u*math.sin(beta/2))*u
            Fc=(C-(math.sin(beta/2)-beta/2*math.cos(beta/2))*f)/((beta/2+u)*math.sin(beta/2)-u*beta/2*math.cos(beta/2))
            ff=(z1*f+z2*C/z4+z5*Fc)/(1-u*math.sin(beta/2)+z2*z3/z4)
    if f>=8000:
        ff=1.1*ff*((f/ff)**0.85)
    bendfriction=ff-f
    return ff,bendfriction


def straight(rou,A,x,u):      #rou是密度，A是截面积，x是长度，u是摩擦系数
    strafriction=rou*A*x*u*9.8*1000
    if x<=0:
        strafriction=0
    return strafriction

def youxiaoL(b1,b2,x,r0,rou,h,b):
    if (b1>=180)&(b2>=180) or (b1<=180)&(b2<=180):
        x=x
    else:
        x=x-4*(8*r0*Ej/rou/9.8/h/b)**0.25/5.6
    if x<0:
        x=0
    return x
        
workbook=xlrd.open_workbook(r'直管弯头输入参数-mm-度.xlsx')
sheet=workbook.sheet_by_index(1)
row = sheet.nrows-1
col = 2
xb=np.zeros((row,col))
for i in range(row):
    for j in range(col):
        if type(sheet.cell(i+1,j+1).value)==type("abc"):
            xb[i][j]=-1
        else:
            xb[i][j]=sheet.cell(i+1,j+1).value
        
Ff=np.zeros((row,3))  
def F(E,R0,r0,b,h,f,rou,u):
    A=b*h
    Ff[0][0]=straight(rou,A,xb[0][0],u)+f
    mum=0
    for i in range(row-1):
        bendi=bend(E,R0,r0,b,h,xb[i][1],Ff[i][0],u)[0]
        Ff[i][1]=bendi
        if xb[i+1][1]==-1:
            L=xb[i+1][0]
        else:
            L=youxiaoL(xb[i][1], xb[i+1][1], xb[i+1][0],r0,rou,h,b)
        Ff[i+1][0]=bendi+straight(rou,A,L,u)
        mum=mum+xb[i][0]+1995*0.4
        Ff[i][2]=mum
    Ff[row-1][2]=Ff[row-2][2]+xb[row-1][0]
    Ff[:,2]=10*Ff[:,2]
    return Ff
f=sheet.cell(1,3).value
miu=sheet.cell(1,4).value
FFFFF=F(1500,1270,127,190,20,f,0.00000000091,miu)/10000
workbook1=openpyxl.Workbook()

booksheet1=workbook1.create_sheet('牵引力')
for i in range(1,row+1):
    for j in range(1,4):
        booksheet1.cell(i,j).value=FFFFF[i-1][j-1]
 
workbook1.save('多弯牵引力计算-吨.xlsx')